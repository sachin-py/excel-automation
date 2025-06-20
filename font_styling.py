"""
📄 Excel Font Styling with openpyxl

This script opens an existing Excel file (`test_font.xlsx`) and applies different 
font styles to specific cells:

- Cell A1 → Bold text
- Cell B2 → Italic text
- Cell C3 → Underlined text

Then, it saves the styled workbook.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/pattern_programs/
"""

from openpyxl import load_workbook            # To load existing workbook
from openpyxl.styles import Font              # To apply font styles

# Load the workbook
workbook = load_workbook('test_font.xlsx')

# Access the first worksheet
sheet = workbook.worksheets[0]

# Apply bold font to cell A1
cell_a1 = sheet.cell(row=1, column=1)
font_a1 = Font(bold=True)
cell_a1.font = font_a1

# Apply italic font to cell B2
cell_b2 = sheet.cell(row=2, column=2)
font_b2 = Font(italic=True)
cell_b2.font = font_b2

# Apply single underline to cell C3 (note: column=2 is B, not C)
cell_c3 = sheet.cell(row=3, column=2)
font_c3 = Font(underline='single')
cell_c3.font = font_c3

# Save the updated workbook
workbook.save('test_font.xlsx')

# Close the workbook
workbook.close()
