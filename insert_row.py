"""
Insert a New Row into an Excel Sheet using openpyxl

This script inserts a new row at position 6 in an existing Excel file (`sample.xlsx`)
and fills it with the values: ['new1', 'new2', 'new3', 'new4', 'new5']

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook  # For loading and modifying Excel workbooks

# Load the workbook
workbook = load_workbook('sample.xlsx')

# Access the target sheet
sheet = workbook['Sheet']

# Define the row number where new data will be inserted
row_to_insert = 6

# Insert an empty row at the specified position
sheet.insert_rows(row_to_insert)

# Define content to add to the inserted row
content_to_add = ['new1', 'new2', 'new3', 'new4', 'new5']

# Populate the inserted row with the defined values
for col, value in enumerate(content_to_add, start=1):
    sheet.cell(row=row_to_insert, column=col, value=value)

# Save the changes
workbook.save('sample.xlsx')

# Close the workbook
workbook.close()
