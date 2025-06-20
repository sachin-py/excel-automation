"""
Excel Row Reader

This script reads and prints the values of cells in a specific row 
(from column 1 to 4) of an Excel (.xlsx) sheet using `openpyxl`.

Row selected: 2  
Columns selected: A to D (1 to 4)

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/pattern_programs/
"""

# Import the load_workbook method from openpyxl
from openpyxl import load_workbook

# Load the workbook named 'test.xlsx'
workbook = load_workbook(filename='test.xlsx')

# Activate the default worksheet
sheet1 = workbook.active

# Define the row number you want to read
rowNo = 2

# Iterate over the specified row (rowNo) from column 1 (A) to column 4 (D)
for row in sheet1.iter_rows(min_row=rowNo, max_row=rowNo, min_col=1, max_col=4):
    for cell in row:
        print(cell.value)  # Print the value of each cell in the row

# Close the workbook after reading
workbook.close()
