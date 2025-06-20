"""
📄 Excel Multiple Sheets Reader

This script demonstrates how to:
1. Access multiple sheets in an Excel workbook by their names.
2. Access the same sheets using the `worksheets` list (by index).
3. Print specific cell values from each sheet.

Sheet names used:
- 'Sheet1'
- 'this is second sheet'
- 'Sheet3'

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/pattern_programs/
"""

from openpyxl import load_workbook  # Import to load and work with Excel files

# Load the workbook
workbook = load_workbook(filename='test.xlsx')

# Access sheets by name
sheet1 = workbook['Sheet1']
sheet2 = workbook['this is second sheet']
sheet3 = workbook['Sheet3']

# Print value from cell B2 (row=2, column=2) in each named sheet
print(sheet1.cell(row=2, column=2).value)
print(sheet2.cell(row=2, column=2).value)
print(sheet3.cell(row=2, column=2).value)

# Access the same sheets using index from workbook.worksheets list
sheet1_1 = workbook.worksheets[0]
sheet2_1 = workbook.worksheets[1]
sheet3_1 = workbook.worksheets[2]

# Print specific cell values using the indexed sheet references
print(sheet1_1.cell(row=2, column=2).value)  # B2 from Sheet1
print(sheet2_1.cell(row=2, column=3).value)  # C2 from Sheet2
print(sheet3_1.cell(row=1, column=2).value)  # B1 from Sheet3

# Close the workbook after use
workbook.close()
