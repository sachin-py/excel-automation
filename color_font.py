"""
Excel Font Color and Background Fill Styling

This script opens an existing Excel file (`font_color.xlsx`) and applies:
- Font colors (red, blue)
- Background fills (green)
- Font + fill combinations

Affected Cells:
- A7 → Red text
- B7 → Default font, green background
- C7 → Blue text with green background

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

# For loading Excel files
from openpyxl import load_workbook
# For font and background styles
from openpyxl.styles import Font, PatternFill

# Load the workbook
workbook = load_workbook('font_color.xlsx')

# Access the specific sheet by name
sheet = workbook['Sheet']

# --- Style A7: Red font only ---
cell_a7 = sheet.cell(row=7, column=1, value='I am red')
font_a7 = Font(color='FF0000')  # Red in RGB
cell_a7.font = font_a7

# --- Style B7: Green background fill, default font ---
cell_b7 = sheet.cell(
    row=7, column=2, value='font is different bg is different')
yellow_fill = PatternFill(start_color='00FF00',
                          end_color='00FF00', fill_type='solid')
cell_b7.fill = yellow_fill

# --- Style C7: Blue font with green background fill ---
cell_c7 = sheet.cell(row=7, column=3, value='mixed')
font_c7 = Font(color='0000FF')  # Blue font
cell_c7.font = font_c7
green_fill = PatternFill(start_color='00FF00',
                         end_color='00FF00', fill_type='solid')
cell_c7.fill = green_fill

# Save changes
workbook.save('font_color.xlsx')

# Close workbook
workbook.close()
