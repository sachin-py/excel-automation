"""
🗑️ Delete a Column from an Excel Sheet using openpyxl

This script deletes column 2 (i.e., column B) from an existing Excel file (`sample.xlsx`)  
and prints the number of columns before and after deletion.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook  # For working with Excel files

# Load the workbook
workbook = load_workbook('sample.xlsx')

# Access the sheet by name
sheet = workbook['Sheet']

# Specify which column to delete (column B = 2)
column_to_del = 2

# Print column count before deletion
print('Number of columns before deletion:', sheet.max_column)

# Delete the specified column
sheet.delete_cols(column_to_del)

# Print column count after deletion
print('Number of columns after deletion:', sheet.max_column)

# Save the modified workbook
workbook.save('sample.xlsx')

# Close the workbook
workbook.close()
