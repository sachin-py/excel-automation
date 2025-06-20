"""
Safe Extraction of Excel Cell Fill Color

Reads the background color of a cell and prints it in RGB hex (if available).
Handles cases where color is not defined as plain RGB.

Author: Sachin Kumar  
GitHub: https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook

# Load workbook and select sheet
workbook = load_workbook('font_color.xlsx')
sheet = workbook['Sheet']

# Target cell (D7 = row 7, col 4)
cell = sheet.cell(row=7, column=4)
fill = cell.fill

# Check if the fill is solid
if fill.fill_type == 'solid':
    start_color = fill.start_color

    # --- Case 1: RGB Color ---
    if start_color.type == "rgb" and start_color.rgb:
        print("Fill Color (RGB):", start_color.rgb)  # Example: FF00FF00

    # --- Case 2: Theme Color (used by Excel default palette) ---
    elif start_color.type == "theme":
        print(
            f"Theme Color (index): {start_color.theme} → Not a direct RGB value")

    # --- Case 3: Indexed Color (older Excel palette) ---
    elif start_color.type == "indexed":
        print(
            f"Indexed Color (index): {start_color.indexed} → Not a direct RGB value")

    else:
        print("Unknown or unsupported color type:", start_color.type)
else:
    print("Cell does not have a solid fill.")
