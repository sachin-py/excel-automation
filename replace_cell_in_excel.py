"""
📄 Excel Dynamic Cell Update

This script opens an existing Excel file (`test2.xlsx`), updates a cell at a 
specific row and column using dynamic variables, and saves the file.

- Updates cell at Row 2, Column 3 (i.e., C2) with the value: "I am a new value"

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/pattern_programs/
"""

from openpyxl import load_workbook  # Import to load and modify Excel files

# Load the existing Excel workbook
workbook = load_workbook(filename='test2.xlsx')

# Get the active worksheet
sheet = workbook.active

# Define the target row and column
row_id = 2
column_id = 3

# Define the new value to write
new_value = 'I am a new value'

# Write the value to the specified cell (C2)
sheet.cell(row=row_id, column=column_id, value=new_value)

# Save the changes back to the workbook
workbook.save('test2.xlsx')

# Close the workbook
workbook.close()
