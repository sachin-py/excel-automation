"""
Excel Column Reader

This script reads and prints the values of cells in a specific column 
(from row 1 to 5) of an Excel (.xlsx) sheet using `openpyxl`.

Column selected: A (1st column)  
Rows selected: 1 to 5

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/pattern_programs/
"""

from openpyxl import load_workbook  # Import to load Excel workbooks

# Load the workbook named 'test.xlsx'
workbook = load_workbook(filename='test.xlsx')

# Get the default active sheet
sheet1 = workbook.active

# Iterate through column A (1st column), from row 1 to 5
for cell in sheet1.iter_cols(min_col=1, max_col=1, min_row=1, max_row=5):
    for col_cell in cell:
        print(col_cell.value)  # Print the value of each cell in column A

# Close the workbook after use
workbook.close()
