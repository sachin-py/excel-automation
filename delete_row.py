"""
Delete a Row from an Excel Sheet using openpyxl

This script opens an existing Excel file (`sample.xlsx`) and deletes a specific row (row 6).
It prints the number of rows before and after deletion.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook  # For loading Excel workbooks

# Load the workbook
worbook = load_workbook('sample.xlsx')
sheet = worbook['Sheet']  # Access the specific sheet

# Print the total number of rows before deletion
print('Number of rows before deletion:', sheet.max_row)

# Define which row to delete
row_to_delete = 6
sheet.delete_rows(row_to_delete)

# Print the total number of rows after deletion
print('Number of rows after deletion:', sheet.max_row)

# Save the workbook with changes
worbook.save('sample.xlsx')
worbook.close()
