"""
✅ Filter Excel Data using pandas and write result with openpyxl

This script:
1. Reads an Excel file (`pass_fail.xlsx`) using pandas.
2. Filters rows where the value in the 'Status' column is 'Pass'.
3. Creates a new sheet named 'Filtered_Data' in the same Excel file.
4. Writes the filtered data (including headers) into that new sheet.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

import pandas as pd
from openpyxl import load_workbook

# Load the Excel file into pandas
worbook = pd.ExcelFile('pass_fail.xlsx')
df = worbook.parse('Sheet')  # Read data from the sheet named 'Sheet'

# Extract the 'Status' column and define filter condition
status_column = df['Status']
filter_criteria = 'Pass'

# Create a boolean mask where status is 'Pass'
is_filtered = status_column == filter_criteria

# Apply the filter to get only 'Pass' rows
filtered_df = df[is_filtered]

# Load the workbook using openpyxl to modify it
workbook = load_workbook('pass_fail.xlsx')

# Create a new sheet where the filtered data will be written
new_sheet = 'Filtered_Data'
filtered_sheet = workbook.create_sheet(title=new_sheet)

# Prepare data (headers + rows) to copy into the new sheet
data_to_copy = [filtered_df.columns.tolist()] + filtered_df.values.tolist()

# Print the data to verify
print(data_to_copy)

# Write each row to the new sheet
for row_data in data_to_copy:
    filtered_sheet.append(row_data)

# Save the workbook with the new sheet
workbook.save('pass_fail.xlsx')
workbook.close()
