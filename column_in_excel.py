"""
Write Data to a Specific Column in Excel using openpyxl

This script opens an existing Excel file (`test2.xlsx`) and writes a list of values  
vertically into a specific column (column F, i.e., column 6).  
Each entry is written to a new row starting from row 1.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook  # For loading and editing Excel files

# Load the workbook and select the active sheet
workbook = load_workbook(filename='test2.xlsx')
sheet = workbook.active

# List of values to write into column 6 (i.e., column F)
data_to_write = ['col1', 'col2', 'col3', 'col4', 'col5', 'col6', 'col7']
col_to_write = 6  # Column F

# Write each value in the list to a new row in the specified column
for row_id, entry in enumerate(data_to_write, start=1):
    sheet.cell(row=row_id, column=col_to_write, value=entry)

# Save and close the workbook
workbook.save('test2.xlsx')
workbook.close()
