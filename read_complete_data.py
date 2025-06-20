"""
Excel Full Sheet Reader

This script reads and prints **all cell values** from the entire active sheet 
of an Excel (.xlsx) file using the `openpyxl` library.

It goes row by row, and prints each cell's value sequentially.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/pattern_programs/
"""

from openpyxl import load_workbook  # Import to handle Excel files

# Load the Excel workbook named 'test.xlsx'
workbook = load_workbook(filename='test.xlsx')

# Get the default active worksheet
sheet1 = workbook.active

# Iterate through all rows and cells in the sheet
for row in sheet1.iter_rows():
    for data in row:
        print(data.value)  # Print each cell's value

# Optionally close the workbook (not mandatory for read-only use)
workbook.close()
