"""
Insert a Column into an Excel Sheet using openpyxl

This script inserts a new column at position 2 (i.e., column B) into an existing Excel file (`sample.xlsx`)  
and fills it with the values: [1, 2, 3, 4, 5]

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook  # For working with Excel files

# Load the workbook and target sheet
workbook = load_workbook('sample.xlsx')
sheet = workbook['Sheet']

# Define which column to insert (column B = 2)
col_to_insert = 2

# Insert an empty column at the specified index
sheet.insert_cols(col_to_insert)

# Values to populate in the inserted column
column_data = [1, 2, 3, 4, 5]

# Write values to the new column
for row, value in enumerate(column_data, start=1):
    sheet.cell(row=row, column=col_to_insert, value=value)

# Save the updated workbook
workbook.save('sample.xlsx')

# Close the workbook
workbook.close()
