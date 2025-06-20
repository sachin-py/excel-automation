"""
Create a Pie Chart in Excel using openpyxl

This script reads student subject marks from an Excel file (`marks.xlsx`) and creates a pie chart.
The pie chart displays:
- Categories from column A (subject names)
- Values from column B (marks)
- It places the chart starting at cell E5

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook                # To load the Excel workbook
# For creating pie chart and referencing data
from openpyxl.chart import PieChart, Reference

# Load the workbook and the specific sheet
workbook = load_workbook('marks.xlsx')
sheet = workbook['Sheet']

# Create a new PieChart object
piechart = PieChart()
piechart.title = 'Subject Marks of Student'

# Define data range (actual values, excluding header)
data = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)

# Define category labels (e.g., subject names)
label = Reference(sheet, min_col=1, max_col=1,
                  min_row=2, max_row=sheet.max_row)

# Include the header as data point title (first row)
data_points = Reference(sheet, min_col=2, max_col=2,
                        min_row=1, max_row=sheet.max_row)

# Add data to the chart (titles_from_data=True picks header as label)
piechart.add_data(data_points, titles_from_data=True)

# Set the labels (categories) for each pie slice
piechart.set_categories(label)

# Add the chart to the sheet at location E5
sheet.add_chart(piechart, 'E5')

# Save the updated workbook
workbook.save('marks.xlsx')
workbook.close()
