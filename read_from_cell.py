"""
Excel Cell Value Reader

This script reads and prints values from specific cells in an Excel (.xlsx) file
using the `openpyxl` library.

Cells accessed:
- B2 (Row 2, Column 2)
- A1 (Row 1, Column 1)
- A100 (Row 100, Column 1)

Author : Sachin Kumar
GitHub : https://github.com/sachin-py/pattern_programs/
"""

# Import the load_workbook method to read Excel files
from openpyxl import load_workbook

# Load the Excel workbook named 'test.xlsx'
workbook = load_workbook(filename='test.xlsx')

# Get the default active worksheet from the workbook
sheet1 = workbook.active

# Access and print the value in cell B2 (Row 2, Column 2)
value_in_cell_b2 = sheet1.cell(row=2, column=2).value
print(value_in_cell_b2)

# Access and print the value in cell A1 (Row 1, Column 1)
value_in_cell_A1 = sheet1.cell(row=1, column=1).value
print(value_in_cell_A1)

# Access and print the value in cell A100 (Row 100, Column 1)
value_in_cell_A100 = sheet1.cell(row=100, column=1).value
print(value_in_cell_A100)
