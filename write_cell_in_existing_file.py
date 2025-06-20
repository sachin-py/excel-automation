"""
Excel Cell Update

This script opens an existing Excel file (test2.xlsx) and updates specific cells:
- Writes "I am at D6" to cell D6 (row 6, column 4)
- Writes "New Value" to cell B7 (row 7, column 2)

After modification, it saves the changes back to the same file.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/pattern_programs/
"""

from openpyxl import load_workbook  # Import to load and edit existing Excel files

# Load the existing Excel workbook
workbook = load_workbook(filename='test2.xlsx')

# Get the active sheet
sheet = workbook.active

# Write 'I am at D6' to cell D6 (Row 6, Column 4)
sheet.cell(row=6, column=4, value='I am at D6')

# Write 'New Value' to cell B7 (Row 7, Column 2)
sheet.cell(row=7, column=2, value='New Value')

# Save the workbook with the changes
workbook.save('test2.xlsx')

# Close the workbook
workbook.close()
