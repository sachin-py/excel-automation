"""
Excel Font Styling in a Row

This script opens an existing Excel file (`test_font.xlsx`) and applies different 
font styles to cells in the **first row** of the first worksheet:

- A1 → Bold
- B1 → Italic
- C1 → Underlined (single)

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import load_workbook          # For loading existing Excel file
from openpyxl.styles import Font            # For styling fonts in cells

# Load the existing workbook
workbook = load_workbook('test_font.xlsx')

# Access the first worksheet
sheet = workbook.worksheets[0]

# --- Apply Bold to cell A1 ---
cell_a1_bold = sheet.cell(row=1, column=1)
font_a1 = Font(bold=True)
cell_a1_bold.font = font_a1

# --- Apply Italic to cell B1 ---
cell_b1_italic = sheet.cell(row=1, column=2)
font_b1 = Font(italic=True)
cell_b1_italic.font = font_b1

# --- Apply Underline to cell C1 ---
# Possible values: 'single', 'double', 'singleAccounting', 'doubleAccounting'
cell_c1_underline = sheet.cell(row=1, column=3)
font_c1 = Font(underline='single')
cell_c1_underline.font = font_c1

# Save the updated workbook
workbook.save('test_font.xlsx')

# Close the workbook
workbook.close()
