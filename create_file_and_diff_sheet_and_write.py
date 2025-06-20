"""
Excel Workbook with Multiple Sheets

This script creates a new Excel workbook with three sheets:
- Default sheet (Sheet)
- Sheet2
- Sheet3

It writes different values into:
- Cell A1 of Sheet1
- Cell B2 of Sheet2
- Cell C3 of Sheet3

Then, it saves the workbook as mynewfile.xlsx.

Author : Sachin Kumar  
GitHub : https://github.com/sachin-py/excel-automation
"""

from openpyxl import Workbook  # Import to create a new Excel workbook

# Create a new workbook
workbook = Workbook()

# Get the default active sheet (Sheet1)
sheet = workbook.active

# Create two more sheets
sheet2 = workbook.create_sheet('Sheet2')
sheet3 = workbook.create_sheet('Sheet3')

# Write data to each sheet at specific cells
sheet.cell(row=1, column=1, value='I am in sheet1')   # A1 in Sheet1
sheet2.cell(row=2, column=2, value='I am in sheet2')  # B2 in Sheet2
sheet3.cell(row=3, column=3, value='I am in sheet3')  # C3 in Sheet3

# Save the workbook to a new file
workbook.save('mynewfile.xlsx')

# Close the workbook
workbook.close()
